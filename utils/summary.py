from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
import requests
import asyncio
import time
from openpyxl import load_workbook

class Summary():
    def __init__(self, file, sheet, i = 0, last = None,  delay = None):
        self.last = last
        self.delay = delay
        self.wb = load_workbook(file)
        self.ws = self.wb[sheet]
        self.chrome_options = Options()
        self.chrome_options.add_experimental_option("detach", True)
        self.chrome_options.add_argument('--headless')
        self.chrome_options.add_argument('--no-sandbox')
        self.chrome_options.add_argument('--disable-dev-shm-usage')
        self.browser = webdriver.Chrome(self.chrome_options)
        self.i = i

    async def track_svo(self, awb_blank = None, awb_num = None, airport = None):
        url = 'https://www.moscow-cargo.com/'
        if requests.get(url).status_code != 200:
            return 'В данный момент трекинг недоступен. Попробуйте позднее'
        awb_blank = self.ws[f"B{self.i}"].value[0:3]
        awb_num = self.ws[f"B{self.i}"].value[4:12]
        self.browser.get(url)

        awb_blank_elem = self.browser.find_element(By.CSS_SELECTOR, '[class = "awb-prefix awb-part"]')
        awb_num_elem = self.browser.find_element(By.CSS_SELECTOR, '[class = "awb-number awb-part"]')

        awb_blank_elem.send_keys(awb_blank)
        awb_num_elem.send_keys(awb_num + Keys.RETURN)
        time.sleep(self.delay)
        try:
            status_els = self.browser.find_elements(By.CSS_SELECTOR, '[id = "statusbody"]')
            cargo_info = status_els[0].find_elements(By.TAG_NAME, "tr")[0].find_elements(By.TAG_NAME, "td")
            info = {}
            info['awb'] = cargo_info[0].text
            info['date'] = cargo_info[7].find_elements(By.TAG_NAME, "span")[1].text
            info['cw'] = cargo_info[3].find_elements(By.TAG_NAME, "div")[0].find_elements(By.TAG_NAME, "span")[1].text
            info['vol'] = cargo_info[4].text
            self.ws[f"A{self.i}"].value = info['date']
            self.ws[f"E{self.i}"].value = float(info['cw'])
            self.ws[f"F{self.i}"].value = float(info['vol'])
            self.i = self.i + 1
            return info 
        except Exception as e:
            print(e)
            self.i = self.i + 1
            return 'Не удалось найти накладную с номером ' + '<code>' + awb_blank + '-' + awb_num + '</code>'
        
async def main():
    tr = Summary(file = '/Users/glebkuimov/Desktop/tesis tr/awbs_tr.xlsx', sheet= "AWB", i = 159, last = 188, delay=5)
    while(tr.i <= tr.last):
        try:
            print(tr.i)
            l = await tr.track_svo()
            print(l)
            tr.wb.close()
            tr.wb.save("/Users/glebkuimov/Desktop/tesis tr/awbs_tr.xlsx")
        except Exception as e:
            print(e)
            pass
    tr.wb.close()
    tr.wb.save("/Users/glebkuimov/Desktop/tesis tr/awbs_tr.xlsx")
asyncio.run(main())